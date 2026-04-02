import tkinter as tk
from tkinter import ttk, filedialog, messagebox
import json
import os
import openpyxl
from openpyxl.styles import PatternFill

BASE_DIR = os.getcwd()
CONFIG_PATH = os.path.join(BASE_DIR, "matching_config.json")
RED_FILL = PatternFill(start_color="FF0000", end_color="FF0000", fill_type="solid")
GREEN_FILL = PatternFill(start_color="00FF00", end_color="00FF00", fill_type="solid")

SHIP_TYPE_LABELS = {
    "free": "free(택배비포함)",
    "fixed": "fixed(고정택배비)",
    "variable": "variable(조건별택배비)",
    "logen_calc": "logen_calc(로젠택배계산)",
}
SHIP_TYPE_KEYS = {v: k for k, v in SHIP_TYPE_LABELS.items()}

DEFAULT_CONFIG = {
    "settings": {
        "default_manager": "온라인",
        "excel_columns": {
            "order_date": 0, "name": 1, "address": 2, "phone": 3,
            "product": 4, "option": 5, "quantity": 6, "zipcode": 7,
            "message": 8
        }
    },
    "vendors": {},
    "excluded_products": []
}


def load_config():
    if os.path.exists(CONFIG_PATH):
        with open(CONFIG_PATH, "r", encoding="utf-8") as f:
            return json.load(f)
    return DEFAULT_CONFIG.copy()


def save_config(config):
    with open(CONFIG_PATH, "w", encoding="utf-8") as f:
        json.dump(config, f, ensure_ascii=False, indent=2)


class App(tk.Tk):
    def __init__(self):
        super().__init__()
        self.title("더해피 발주 자동분류")
        self.geometry("1100x700")
        ico = os.path.join(os.getcwd(), "softcat2.ico")
        if os.path.exists(ico):
            self.iconbitmap(ico)
        self.config_data = load_config()
        self.orders = []
        self._sort_states = {}  # tree_id -> {col: 'asc'|'desc'|None}

        nb = ttk.Notebook(self)
        nb.pack(fill="both", expand=True, padx=5, pady=5)

        self.tab_order = ttk.Frame(nb)
        self.tab_matching = ttk.Frame(nb)
        self.tab_vendor = ttk.Frame(nb)
        self.tab_settings = ttk.Frame(nb)

        nb.add(self.tab_order, text="발주 처리")
        nb.add(self.tab_matching, text="매칭 관리")
        nb.add(self.tab_vendor, text="업체 관리")
        nb.add(self.tab_settings, text="설정")

        self._build_order_tab()
        self._build_matching_tab()
        self._build_vendor_tab()
        self._build_settings_tab()

    def _make_sortable(self, tree):
        """Treeview 헤더 클릭 시 오름차순→내림차순→원래순서 토글"""
        tid = id(tree)
        self._sort_states[tid] = {}
        for col in tree["columns"]:
            tree.heading(col, command=lambda c=col, t=tree, ti=tid: self._sort_column(t, ti, c))

    def _sort_column(self, tree, tid, col):
        state = self._sort_states[tid].get(col)
        # 다른 컬럼 정렬 해제 표시 초기화
        for c in tree["columns"]:
            tree.heading(c, text=c.replace(" ▲", "").replace(" ▼", ""))
        if state is None:
            reverse = False
            self._sort_states[tid] = {col: "asc"}
            suffix = " ▲"
        elif state == "asc":
            reverse = True
            self._sort_states[tid] = {col: "desc"}
            suffix = " ▼"
        else:
            self._sort_states[tid] = {}
            # 원래 순서 복원 — 각 트리뷰의 refresh 호출
            if tree is self.order_tree:
                self._refresh_order_tree()
            elif tree is self.match_tree:
                self._refresh_match_tree()
            elif tree is self.vendor_tree:
                self._refresh_vendor_tree()
            return

        items = [(tree.item(k, "values"), k) for k in tree.get_children()]
        col_idx = list(tree["columns"]).index(col)
        def sort_key(x):
            v = x[0][col_idx]
            try:
                return (0, float(v))
            except (ValueError, TypeError):
                return (1, str(v))
        items.sort(key=sort_key, reverse=reverse)
        for i, (_, k) in enumerate(items):
            tree.move(k, "", i)
        label = col + suffix
        tree.heading(col, text=label)

    # ── 발주 처리 탭 ──
    def _build_order_tab(self):
        top = ttk.Frame(self.tab_order)
        top.pack(fill="x", padx=5, pady=5)

        ttk.Button(top, text="엑셀 파일 열기", command=self._load_excel).pack(side="left")
        self.lbl_file = ttk.Label(top, text="파일을 선택하세요")
        self.lbl_file.pack(side="left", padx=10)

        ttk.Label(top, text="검색:").pack(side="left", padx=(20, 0))
        self.order_search_var = tk.StringVar()
        self.order_search_var.trace_add("write", lambda *_: self._refresh_order_tree())
        ttk.Entry(top, textvariable=self.order_search_var, width=20).pack(side="left", padx=5)

        # 주문 테이블
        tree_frame = ttk.Frame(self.tab_order)
        tree_frame.pack(fill="both", expand=True, padx=5)
        cols = ("주문일시", "수취인", "주소", "전화번호", "상품명", "옵션이름", "수량", "배송메세지", "매칭업체", "품목")
        self.order_tree = ttk.Treeview(tree_frame, columns=cols, show="headings", height=20)
        vsb = ttk.Scrollbar(tree_frame, orient="vertical", command=self.order_tree.yview)
        hsb = ttk.Scrollbar(tree_frame, orient="horizontal", command=self.order_tree.xview)
        self.order_tree.configure(yscrollcommand=vsb.set, xscrollcommand=hsb.set)
        self.order_tree.grid(row=0, column=0, sticky="nsew")
        vsb.grid(row=0, column=1, sticky="ns")
        hsb.grid(row=1, column=0, sticky="ew")
        tree_frame.grid_rowconfigure(0, weight=1)
        tree_frame.grid_columnconfigure(0, weight=1)
        for c in cols:
            self.order_tree.heading(c, text=c)
            self.order_tree.column(c, width=100 if c not in ("주소",) else 200)

        # 태그 색상
        self.order_tree.tag_configure("unmatched", background="#ffcccc")
        self.order_tree.tag_configure("matched", background="#ccffcc")
        self.order_tree.tag_configure("excluded", background="#dddddd")
        self._make_sortable(self.order_tree)

        bottom = ttk.Frame(self.tab_order)
        bottom.pack(fill="x", padx=5, pady=5)

        self.lbl_status = ttk.Label(bottom, text="")
        self.lbl_status.pack(side="left")
        ttk.Button(bottom, text="선택 삭제", command=self._del_order).pack(side="left", padx=5)
        ttk.Button(bottom, text="전체 삭제", command=self._del_all_orders).pack(side="left", padx=5)
        ttk.Button(bottom, text="자동 매칭 실행", command=self._run_matching).pack(side="right", padx=5)
        ttk.Button(bottom, text="부반장제어 저장", command=self._save_temp_excel).pack(side="right", padx=5)
        ttk.Button(bottom, text="구글시트 전송", command=self._send_to_sheets).pack(side="right", padx=5)

        # 더블클릭 수기 수정
        self.order_tree.bind("<Double-1>", self._on_order_dblclick)

    def _del_order(self):
        sel = self.order_tree.selection()
        if not sel:
            messagebox.showwarning("알림", "삭제할 항목을 선택하세요")
            return
        indices = []
        all_items = self.order_tree.get_children()
        for s in sel:
            indices.append(all_items.index(s))
        for i in sorted(indices, reverse=True):
            del self.orders[i]
        self._refresh_order_tree()
        self.lbl_status.config(text=f"총 {len(self.orders)}건")

    def _del_all_orders(self):
        if not self.orders:
            return
        if messagebox.askyesno("확인", "전체 주문을 삭제하시겠습니까?"):
            self.orders.clear()
            self._refresh_order_tree()
            self.lbl_status.config(text="")

    def _on_order_dblclick(self, event):
        item = self.order_tree.identify_row(event.y)
        col = self.order_tree.identify_column(event.x)
        if not item or not col:
            return
        col_idx = int(col.replace("#", "")) - 1
        cols = ("주문일시", "수취인", "주소", "전화번호", "상품명", "옵션이름", "수량", "배송메세지", "매칭업체", "품목")
        field_map = {0: "date", 1: "name", 2: "address", 3: "phone",
                     4: "product", 5: "option", 6: "quantity", 7: "message"}
        if col_idx not in field_map:
            return
        bbox = self.order_tree.bbox(item, col)
        if not bbox:
            return
        old_val = self.order_tree.item(item, "values")[col_idx]
        entry = ttk.Entry(self.order_tree)
        entry.place(x=bbox[0], y=bbox[1], width=bbox[2], height=bbox[3])
        entry.insert(0, old_val)
        entry.select_range(0, "end")
        entry.focus()

        all_items = self.order_tree.get_children()
        order_idx = all_items.index(item)
        field = field_map[col_idx]

        def _save(e=None):
            val = entry.get()
            if field == "quantity":
                try:
                    val = int(val)
                except ValueError:
                    val = 0
            self.orders[order_idx][field] = val
            entry.destroy()
            self._refresh_order_tree()

        entry.bind("<Return>", _save)
        entry.bind("<FocusOut>", _save)
        entry.bind("<Escape>", lambda e: entry.destroy())

    def _load_excel(self):
        path = filedialog.askopenfilename(filetypes=[("Excel", "*.xlsx")])
        if not path:
            return
        self.lbl_file.config(text=os.path.basename(path))
        from openpyxl.styles.fills import Fill
        _orig_init = Fill.__init__
        if _orig_init is object.__init__:
            Fill.__init__ = lambda self, *a, **kw: None
        try:
            wb = openpyxl.load_workbook(path)
        finally:
            Fill.__init__ = _orig_init
        ws = wb.active
        self.orders = []
        col = self.config_data["settings"]["excel_columns"]
        for row in ws.iter_rows(min_row=2, max_row=ws.max_row, values_only=True):
            if not any(row):
                continue
            self.orders.append({
                "date": str(row[col["order_date"]] or ""),
                "name": str(row[col["name"]] or ""),
                "address": str(row[col["address"]] or ""),
                "phone": str(row[col["phone"]] or ""),
                "product": str(row[col["product"]] or ""),
                "option": str(row[col["option"]] or ""),
                "quantity": row[col["quantity"]] or 0,
                "zipcode": str(row[col["zipcode"]] or "") if col.get("zipcode") is not None else "",
                "message": str(row[col["message"]] or ""),
                "vendor": None,
                "vendor_id": None,
                "item_name": None,
                "shipping_type": None,
                "shipping_fee": None,
                "excluded": False,
            })
        self._refresh_order_tree()
        self.lbl_status.config(text=f"총 {len(self.orders)}건 로드됨")

    def _refresh_order_tree(self):
        self.order_tree.delete(*self.order_tree.get_children())
        keyword = self.order_search_var.get().strip().lower() if hasattr(self, "order_search_var") else ""
        for o in self.orders:
            if keyword:
                row_text = " ".join(str(v) for v in [
                    o["date"], o["name"], o["address"], o["phone"],
                    o["product"], o["option"], o["quantity"], o["message"],
                    o["vendor"] or "", o["item_name"] or ""
                ]).lower()
                if keyword not in row_text:
                    continue
            if o["excluded"]:
                tag = "excluded"
            elif o["vendor"]:
                tag = "matched"
            else:
                tag = "unmatched"
            self.order_tree.insert("", "end", values=(
                o["date"], o["name"], o["address"], o["phone"],
                o["product"], o["option"], o["quantity"], o["message"],
                "비대상" if o["excluded"] else (o["vendor"] or "미매칭"),
                o["item_name"] or ""
            ), tags=(tag,))

    def _build_matching_lookup(self):
        """상품명/옵션이름 → (vendor_id, vendor_name, item_name, ship_type, ship_fee, override_qty) 매핑"""
        lookup = {}
        for vid, v in self.config_data.get("vendors", {}).items():
            for keyword, pinfo in v.get("products", {}).items():
                if isinstance(pinfo, dict):
                    lookup[keyword] = (vid, v["name"], pinfo["item_name"],
                                       pinfo.get("shipping_type", "free"),
                                       pinfo.get("shipping_fee"),
                                       pinfo.get("override_qty", 1))
                else:
                    lookup[keyword] = (vid, v["name"], pinfo, "free", None, 1)
        return lookup

    def _run_matching(self):
        if not self.orders:
            messagebox.showwarning("알림", "엑셀을 먼저 로드하세요")
            return
        lookup = self._build_matching_lookup()
        excluded = set(self.config_data.get("excluded_products", []))
        matched = 0
        excluded_cnt = 0
        for o in self.orders:
            key = o["option"] if o["option"] else o["product"]
            if key in excluded:
                o["vendor"] = None
                o["item_name"] = None
                o["shipping_type"] = None
                o["shipping_fee"] = None
                o["excluded"] = True
                excluded_cnt += 1
            elif key in lookup:
                vid, o["vendor"], o["item_name"], o["shipping_type"], o["shipping_fee"], oq = lookup[key]
                o["vendor_id"] = vid
                if oq and oq > 1:
                    o["quantity"] = oq
                o["excluded"] = False
                matched += 1
            else:
                o["vendor"] = None
                o["vendor_id"] = None
                o["item_name"] = None
                o["shipping_type"] = None
                o["shipping_fee"] = None
                o["excluded"] = False
        self._refresh_order_tree()
        unmatched = len(self.orders) - matched - excluded_cnt
        msg = f"매칭: {matched}건 성공"
        if excluded_cnt:
            msg += f", {excluded_cnt}건 비대상"
        if unmatched:
            msg += f", {unmatched}건 미매칭"
        self.lbl_status.config(text=msg)
        if unmatched:
            messagebox.showwarning("미매칭 경고", f"{unmatched}건의 미매칭 상품이 있습니다.\n매칭 관리 탭에서 등록해주세요.")

    def _save_temp_excel(self):
        if not self.orders:
            return
        path = filedialog.asksaveasfilename(defaultextension=".xlsx", initialfile="부반장제어_임시.xlsx")
        if not path:
            return

        # 원본 부반장제어 파일 복사해서 사용
        src = os.path.join(BASE_DIR, "부반장제어파일 계속수정.xlsx")
        if os.path.exists(src):
            wb = openpyxl.load_workbook(src)
        else:
            messagebox.showerror("오류", "원본 부반장제어 파일이 없습니다.\n'부반장제어파일 계속수정.xlsx'를 프로그램 폴더에 넣어주세요.")
            return

        # 업체별(vendor_id) 분류
        vendors_data = {}
        unmatched = []
        for o in self.orders:
            if o["excluded"]:
                continue
            vid = o.get("vendor_id")
            if vid:
                vendors_data.setdefault(vid, []).append(o)
            else:
                unmatched.append(o)

        # 미매칭 시트를 맨 앞에 추가
        ws_um = wb.create_sheet("미매칭", 0)
        ws_um.append(["수취인", "주소", "전화번호", "상품명", "옵션이름", "수량", "배송메세지"])
        for o in unmatched:
            ws_um.append([o["name"], o["address"], o["phone"],
                          o["product"], o["option"], o["quantity"], o["message"]])

        # 각 업체 시트에 데이터 입력
        for vid, rows in vendors_data.items():
            v = self.config_data["vendors"].get(vid, {})
            sheet_name = v.get("sheet_name", "")
            if sheet_name not in wb.sheetnames:
                continue
            ws = wb[sheet_name]

            # 2행 헤더에서 컬럼 위치 파악
            headers = [str(ws.cell(row=2, column=c).value or "").strip() for c in range(1, ws.max_column + 1)]

            def find_col(names):
                for n in names:
                    for i, h in enumerate(headers):
                        if h == n:
                            return i + 1  # 1-based
                return None

            c_name = find_col(["상호"])
            c_addr = find_col(["주소"])
            c_phone = find_col(["전화번호"])
            c_item = find_col(["품목"])
            c_qty = find_col(["수량"])
            c_ship = find_col(["택배비"])
            c_zip = find_col(["우편번호"])
            c_mgr = find_col(["담당", "담당자"])
            c_note = find_col(["특이사항"])

            mgr = v.get("default_manager") or self.config_data["settings"].get("default_manager", "온라인")

            # 기존 데이터 아래 빈 행 찾기 (3행부터)
            start_row = 3
            for r in range(3, ws.max_row + 2):
                if all(ws.cell(row=r, column=c).value is None for c in range(1, len(headers) + 1)):
                    start_row = r
                    break

            row_num = start_row
            for o in rows:
                ship_fee = ""
                ship_type = o.get("shipping_type", "free")
                if ship_type in ("fixed", "variable") and o.get("shipping_fee"):
                    raw = str(o["shipping_fee"]).strip().lstrip("'")
                    ship_fee = int(raw) if raw.isdigit() else raw

                if c_name: ws.cell(row=row_num, column=c_name, value=o["name"])
                if c_addr: ws.cell(row=row_num, column=c_addr, value=o["address"])
                if c_phone: ws.cell(row=row_num, column=c_phone, value=o["phone"])
                if c_item: ws.cell(row=row_num, column=c_item, value=o["item_name"] or "")
                if c_qty: ws.cell(row=row_num, column=c_qty, value=o["quantity"])
                if c_ship: ws.cell(row=row_num, column=c_ship, value=ship_fee)
                if c_zip: ws.cell(row=row_num, column=c_zip, value=o.get("zipcode", ""))
                if c_mgr: ws.cell(row=row_num, column=c_mgr, value=mgr)
                if c_note: ws.cell(row=row_num, column=c_note, value=o["message"])

                if ship_type in ("fixed", "variable") and ship_fee and c_ship:
                    ws.cell(row=row_num, column=c_ship).fill = RED_FILL

                row_num += 1

        wb.save(path)
        messagebox.showinfo("저장 완료", f"부반장제어 파일 저장: {path}")

    def _send_to_sheets(self):
        try:
            import gspread
            from google.oauth2.service_account import Credentials
        except ImportError:
            messagebox.showerror("오류", "gspread, google-auth 패키지를 설치해주세요.\npip install gspread google-auth")
            return

        cred_path = os.path.join(BASE_DIR, "credentials.json")
        if not os.path.exists(cred_path):
            messagebox.showerror("오류", f"Google 서비스 계정 인증 파일이 필요합니다.\n{cred_path}")
            return

        matched_orders = [o for o in self.orders if o["vendor"] and not o["excluded"]]
        if not matched_orders:
            messagebox.showwarning("알림", "전송할 매칭된 주문이 없습니다.")
            return

        try:
            scopes = ["https://www.googleapis.com/auth/spreadsheets"]
            creds = Credentials.from_service_account_file(cred_path, scopes=scopes)
            gc = gspread.authorize(creds)

            # 업체별 분류
            vendors_data = {}
            for o in matched_orders:
                vid = o.get("vendor_id")
                vendors_data.setdefault(vid, []).append(o)

            mgr = self.config_data["settings"].get("default_manager", "온라인")
            sent = 0
            for vid, rows in vendors_data.items():
                v = self.config_data["vendors"].get(vid, {})
                url = v.get("google_sheet_url", "")
                if not url:
                    continue
                sh = gc.open_by_url(url)
                try:
                    ws = sh.worksheet("오늘의주문")
                except gspread.WorksheetNotFound:
                    ws = sh.sheet1

                # A열(상호명) 기준으로 300행까지 검사하여 마지막 데이터 행 찾기
                col_a = ws.col_values(1)[:300]
                start_row = 3
                for i in range(len(col_a) - 1, -1, -1):
                    if col_a[i] and str(col_a[i]).strip():
                        start_row = i + 2  # 0-based → 1-based + 다음행
                        break

                new_rows = []
                logen_total = 0
                for o in rows:
                    ship_fee = ""
                    if o.get("shipping_type") in ("fixed", "variable") and o.get("shipping_fee"):
                        raw = str(o["shipping_fee"]).strip().lstrip("'")
                        ship_fee = int(raw) if raw.isdigit() else raw
                    new_rows.append([o["name"], o["address"], o["phone"],
                                     o["item_name"] or "", o["quantity"],
                                     ship_fee, "", mgr, o["message"]])
                    if o.get("shipping_type") == "logen_calc":
                        logen_total += (o["quantity"] if isinstance(o["quantity"], (int, float)) else 1)

                if logen_total > 0:
                    new_rows.append(["", "", "", "로젠택배", logen_total, "", "", mgr, ""])

                if new_rows:
                    ws.update(values=new_rows, range_name=f"A{start_row}")
                    sent += len(rows)

            messagebox.showinfo("전송 완료", f"구글시트에 {sent}건 전송 완료")
        except Exception as e:
            messagebox.showerror("전송 오류", str(e))

    # ── 매칭 관리 탭 ──
    def _build_matching_tab(self):
        top = ttk.Frame(self.tab_matching)
        top.pack(fill="x", padx=5, pady=5)

        ttk.Label(top, text="업체 선택:").pack(side="left")
        self.match_vendor_var = tk.StringVar()
        self.match_vendor_cb = ttk.Combobox(top, textvariable=self.match_vendor_var, state="readonly", width=30)
        self.match_vendor_cb.pack(side="left", padx=5)
        self.match_vendor_cb.bind("<<ComboboxSelected>>", lambda e: self._refresh_match_tree())
        ttk.Button(top, text="새로고침", command=self._refresh_vendor_lists).pack(side="left")

        # 매칭 테이블
        cols = ("상품키워드", "품목명", "수량", "택배비유형", "택배비금액")
        self.match_tree = ttk.Treeview(self.tab_matching, columns=cols, show="headings", height=18)
        for c in cols:
            self.match_tree.heading(c, text=c)
            self.match_tree.column(c, width=60 if c == "수량" else 200)
        self.match_tree.pack(fill="both", expand=True, padx=5)
        self._make_sortable(self.match_tree)

        # 입력 폼
        form = ttk.LabelFrame(self.tab_matching, text="매칭 추가/수정")
        form.pack(fill="x", padx=5, pady=5)

        r = ttk.Frame(form)
        r.pack(fill="x", padx=5, pady=2)
        ttk.Label(r, text="상품키워드(상품명/옵션이름):").pack(side="left")
        self.match_keyword = ttk.Entry(r, width=30)
        self.match_keyword.pack(side="left", padx=5)
        ttk.Label(r, text="품목명:").pack(side="left")
        self.match_item = ttk.Entry(r, width=20)
        self.match_item.pack(side="left", padx=5)
        ttk.Label(r, text="수량:").pack(side="left")
        self.match_qty = ttk.Entry(r, width=5)
        self.match_qty.pack(side="left", padx=5)
        self.match_qty.insert(0, "1")
        ttk.Label(r, text="택배비유형:").pack(side="left")
        self.match_ship_type = ttk.Combobox(r, values=list(SHIP_TYPE_LABELS.values()), width=25, state="readonly")
        self.match_ship_type.pack(side="left", padx=5)
        self.match_ship_type.set(SHIP_TYPE_LABELS["free"])
        ttk.Label(r, text="택배비금액:").pack(side="left")
        self.match_ship_amt = ttk.Entry(r, width=10)
        self.match_ship_amt.pack(side="left", padx=5)

        btns = ttk.Frame(form)
        btns.pack(fill="x", padx=5, pady=2)
        ttk.Button(btns, text="추가", command=self._add_match).pack(side="left", padx=2)
        ttk.Button(btns, text="선택 수정", command=self._update_match).pack(side="left", padx=2)
        ttk.Button(btns, text="선택 삭제", command=self._del_match).pack(side="left", padx=2)

        self.match_tree.bind("<<TreeviewSelect>>", self._on_match_select)
        self._refresh_vendor_lists()

    def _refresh_vendor_lists(self):
        names = sorted([v["name"] for v in self.config_data.get("vendors", {}).values()])
        self.match_vendor_cb["values"] = names
        if hasattr(self, "vendor_tree"):
            self._refresh_vendor_tree()

    def _refresh_match_tree(self):
        self.match_tree.delete(*self.match_tree.get_children())
        vname = self.match_vendor_var.get()
        for vid, v in self.config_data.get("vendors", {}).items():
            if v["name"] == vname:
                for kw, pinfo in sorted(v.get("products", {}).items()):
                    if isinstance(pinfo, dict):
                        stype = pinfo.get("shipping_type", "free")
                        self.match_tree.insert("", "end", values=(
                            kw, pinfo.get("item_name", ""),
                            pinfo.get("override_qty", 1),
                            SHIP_TYPE_LABELS.get(stype, stype),
                            pinfo.get("shipping_fee", "")))
                    else:
                        self.match_tree.insert("", "end", values=(kw, pinfo, 1, SHIP_TYPE_LABELS["free"], ""))
                break

    def _on_match_select(self, event):
        sel = self.match_tree.selection()
        if sel:
            vals = self.match_tree.item(sel[0], "values")
            self.match_keyword.delete(0, "end"); self.match_keyword.insert(0, vals[0])
            self.match_item.delete(0, "end"); self.match_item.insert(0, vals[1])
            self.match_qty.delete(0, "end"); self.match_qty.insert(0, vals[2])
            self.match_ship_type.set(vals[3])
            self.match_ship_amt.delete(0, "end"); self.match_ship_amt.insert(0, vals[4])

    def _find_vendor_id(self, name):
        for vid, v in self.config_data.get("vendors", {}).items():
            if v["name"] == name:
                return vid
        return None

    def _add_match(self):
        vname = self.match_vendor_var.get()
        vid = self._find_vendor_id(vname)
        if not vid:
            messagebox.showwarning("알림", "업체를 선택하세요")
            return
        kw = self.match_keyword.get().strip()
        item = self.match_item.get().strip()
        if not kw or not item:
            messagebox.showwarning("알림", "상품키워드와 품목명을 입력하세요")
            return
        if kw in self.config_data["vendors"][vid].get("products", {}):
            messagebox.showwarning("알림", f"'{kw}' 키워드가 이미 존재합니다. '선택 수정'을 사용하세요.")
            return
        self.config_data["vendors"][vid]["products"][kw] = {
            "item_name": item,
            "override_qty": int(self.match_qty.get() or 1),
            "shipping_type": SHIP_TYPE_KEYS.get(self.match_ship_type.get(), self.match_ship_type.get()),
            "shipping_fee": self.match_ship_amt.get().strip() or None
        }
        save_config(self.config_data)
        self._refresh_match_tree()

    def _update_match(self):
        sel = self.match_tree.selection()
        if not sel:
            messagebox.showwarning("알림", "수정할 항목을 선택하세요")
            return
        old_kw = self.match_tree.item(sel[0], "values")[0]
        vname = self.match_vendor_var.get()
        vid = self._find_vendor_id(vname)
        if not vid:
            return
        new_kw = self.match_keyword.get().strip()
        item = self.match_item.get().strip()
        if not new_kw or not item:
            messagebox.showwarning("알림", "상품키워드와 품목명을 입력하세요")
            return
        # 키워드가 변경된 경우 기존 삭제
        if old_kw != new_kw and old_kw in self.config_data["vendors"][vid]["products"]:
            del self.config_data["vendors"][vid]["products"][old_kw]
        self.config_data["vendors"][vid]["products"][new_kw] = {
            "item_name": item,
            "override_qty": int(self.match_qty.get() or 1),
            "shipping_type": SHIP_TYPE_KEYS.get(self.match_ship_type.get(), self.match_ship_type.get()),
            "shipping_fee": self.match_ship_amt.get().strip() or None
        }
        save_config(self.config_data)
        self._refresh_match_tree()

    def _del_match(self):
        sel = self.match_tree.selection()
        if not sel:
            return
        kw = self.match_tree.item(sel[0], "values")[0]
        vname = self.match_vendor_var.get()
        vid = self._find_vendor_id(vname)
        if vid and kw in self.config_data["vendors"][vid]["products"]:
            del self.config_data["vendors"][vid]["products"][kw]
            save_config(self.config_data)
            self._refresh_match_tree()

    # ── 업체 관리 탭 ──
    def _build_vendor_tab(self):
        cols = ("업체ID", "업체명", "시트이름", "구글시트URL", "담당고정값")
        self.vendor_tree = ttk.Treeview(self.tab_vendor, columns=cols, show="headings", height=18)
        for c in cols:
            self.vendor_tree.heading(c, text=c)
            w = 120 if c != "구글시트URL" else 300
            self.vendor_tree.column(c, width=w)
        self.vendor_tree.pack(fill="both", expand=True, padx=5, pady=5)
        self._make_sortable(self.vendor_tree)

        form = ttk.LabelFrame(self.tab_vendor, text="업체 추가/수정")
        form.pack(fill="x", padx=5, pady=5)

        r1 = ttk.Frame(form); r1.pack(fill="x", padx=5, pady=2)
        ttk.Label(r1, text="업체ID:").pack(side="left")
        self.v_id = ttk.Entry(r1, width=15); self.v_id.pack(side="left", padx=5)
        ttk.Label(r1, text="업체명:").pack(side="left")
        self.v_name = ttk.Entry(r1, width=20); self.v_name.pack(side="left", padx=5)
        ttk.Label(r1, text="시트이름:").pack(side="left")
        self.v_sheet = ttk.Entry(r1, width=25); self.v_sheet.pack(side="left", padx=5)

        r2 = ttk.Frame(form); r2.pack(fill="x", padx=5, pady=2)
        ttk.Label(r2, text="구글시트URL:").pack(side="left")
        self.v_url = ttk.Entry(r2, width=60); self.v_url.pack(side="left", padx=5)
        ttk.Label(r2, text="담당고정값:").pack(side="left")
        self.v_mgr = ttk.Entry(r2, width=10); self.v_mgr.pack(side="left", padx=5)

        btns = ttk.Frame(form); btns.pack(fill="x", padx=5, pady=2)
        ttk.Button(btns, text="추가/수정", command=self._save_vendor).pack(side="left", padx=2)
        ttk.Button(btns, text="선택 삭제", command=self._del_vendor).pack(side="left", padx=2)

        self.vendor_tree.bind("<<TreeviewSelect>>", self._on_vendor_select)
        self._refresh_vendor_tree()

    def _refresh_vendor_tree(self):
        self.vendor_tree.delete(*self.vendor_tree.get_children())
        for vid, v in sorted(self.config_data.get("vendors", {}).items(), key=lambda x: x[1].get("name", "")):
            self.vendor_tree.insert("", "end", values=(
                vid, v.get("name", ""), v.get("sheet_name", ""),
                v.get("google_sheet_url", ""), v.get("default_manager", "")
            ))

    def _on_vendor_select(self, event):
        sel = self.vendor_tree.selection()
        if sel:
            vals = self.vendor_tree.item(sel[0], "values")
            for entry, val in zip([self.v_id, self.v_name, self.v_sheet, self.v_url, self.v_mgr], vals):
                entry.delete(0, "end"); entry.insert(0, val)

    def _save_vendor(self):
        vid = self.v_id.get().strip()
        name = self.v_name.get().strip()
        if not vid or not name:
            messagebox.showwarning("알림", "업체ID와 업체명을 입력하세요")
            return
        if vid not in self.config_data["vendors"]:
            self.config_data["vendors"][vid] = {"products": {}}
        self.config_data["vendors"][vid].update({
            "name": name,
            "sheet_name": self.v_sheet.get().strip(),
            "google_sheet_url": self.v_url.get().strip(),
            "default_manager": self.v_mgr.get().strip() or self.config_data["settings"]["default_manager"]
        })
        save_config(self.config_data)
        self._refresh_vendor_tree()
        self._refresh_vendor_lists()

    def _del_vendor(self):
        sel = self.vendor_tree.selection()
        if not sel:
            return
        vid = self.vendor_tree.item(sel[0], "values")[0]
        if messagebox.askyesno("확인", f"'{vid}' 업체를 삭제하시겠습니까?"):
            self.config_data["vendors"].pop(vid, None)
            save_config(self.config_data)
            self._refresh_vendor_tree()
            self._refresh_vendor_lists()

    # ── 설정 탭 ──
    def _build_settings_tab(self):
        f = ttk.LabelFrame(self.tab_settings, text="기본 설정")
        f.pack(fill="x", padx=10, pady=10)

        r = ttk.Frame(f); r.pack(fill="x", padx=5, pady=5)
        ttk.Label(r, text="담당 고정값 (기본):").pack(side="left")
        self.set_mgr = ttk.Entry(r, width=15)
        self.set_mgr.pack(side="left", padx=5)
        self.set_mgr.insert(0, self.config_data["settings"].get("default_manager", "온라인"))

        ttk.Button(f, text="저장", command=self._save_settings).pack(padx=5, pady=5, anchor="w")

        info = ttk.LabelFrame(self.tab_settings, text="엑셀 컬럼 매핑 (0부터 시작)")
        info.pack(fill="x", padx=10, pady=10)

        self.col_entries = {}
        for i, (key, default) in enumerate(DEFAULT_CONFIG["settings"]["excel_columns"].items()):
            r = ttk.Frame(info); r.pack(fill="x", padx=5, pady=1)
            ttk.Label(r, text=f"{key}:", width=15).pack(side="left")
            e = ttk.Entry(r, width=5)
            e.pack(side="left")
            e.insert(0, str(self.config_data["settings"]["excel_columns"].get(key, default)))
            self.col_entries[key] = e

        # 비대상 상품 관리
        ex_frame = ttk.LabelFrame(self.tab_settings, text="비대상 상품 (발주 제외 품목)")
        ex_frame.pack(fill="x", padx=10, pady=10)

        ex_search_frame = ttk.Frame(ex_frame)
        ex_search_frame.pack(fill="x", padx=5, pady=2)
        ttk.Label(ex_search_frame, text="검색:").pack(side="left")
        self.excluded_search_var = tk.StringVar()
        self.excluded_search_var.trace_add("write", lambda *_: self._refresh_excluded_list())
        ttk.Entry(ex_search_frame, textvariable=self.excluded_search_var, width=30).pack(side="left", padx=5)

        ex_list_frame = ttk.Frame(ex_frame)
        ex_list_frame.pack(fill="x", padx=5, pady=2)
        ex_vsb = ttk.Scrollbar(ex_list_frame, orient="vertical")
        self.excluded_list = tk.Listbox(ex_list_frame, height=6, yscrollcommand=ex_vsb.set)
        ex_vsb.config(command=self.excluded_list.yview)
        self.excluded_list.pack(side="left", fill="x", expand=True)
        ex_vsb.pack(side="right", fill="y")
        self._refresh_excluded_list()

        ex_input = ttk.Frame(ex_frame)
        ex_input.pack(fill="x", padx=5, pady=2)
        self.excluded_entry = ttk.Entry(ex_input, width=30)
        self.excluded_entry.pack(side="left", padx=2)
        ttk.Button(ex_input, text="추가", command=self._add_excluded).pack(side="left", padx=2)
        ttk.Button(ex_input, text="선택 삭제", command=self._del_excluded).pack(side="left", padx=2)

    def _refresh_excluded_list(self):
        self.excluded_list.delete(0, "end")
        keyword = self.excluded_search_var.get().strip().lower() if hasattr(self, "excluded_search_var") else ""
        for item in sorted(self.config_data.get("excluded_products", [])):
            if not keyword or keyword in item.lower():
                self.excluded_list.insert("end", item)

    def _add_excluded(self):
        kw = self.excluded_entry.get().strip()
        if not kw:
            return
        if "excluded_products" not in self.config_data:
            self.config_data["excluded_products"] = []
        if kw not in self.config_data["excluded_products"]:
            self.config_data["excluded_products"].append(kw)
            save_config(self.config_data)
            self._refresh_excluded_list()
        self.excluded_entry.delete(0, "end")

    def _del_excluded(self):
        sel = self.excluded_list.curselection()
        if not sel:
            return
        kw = self.excluded_list.get(sel[0])
        self.config_data.get("excluded_products", []).remove(kw)
        save_config(self.config_data)
        self._refresh_excluded_list()

    def _save_settings(self):
        self.config_data["settings"]["default_manager"] = self.set_mgr.get().strip()
        for key, entry in self.col_entries.items():
            try:
                self.config_data["settings"]["excel_columns"][key] = int(entry.get())
            except ValueError:
                pass
        save_config(self.config_data)
        messagebox.showinfo("저장", "설정이 저장되었습니다.")


if __name__ == "__main__":
    app = App()
    app.mainloop()
